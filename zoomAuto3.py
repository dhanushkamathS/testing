import pyautogui as pya
import time 
from time import sleep
from openpyxl import load_workbook
import sys
import os


# checks if the required input has been given i.e excel file and class number 
def inputChecker():
	try:
		excelFileName = sys.argv[1]
		classNo = int(sys.argv[2])
		return  excelFileName ,classNo 
	except:
		print("error : expects two arguments excelfile and class number") 
		sys.exit()


# checks if give class no is valid
def isClassNoValid(classNo):
	if classNo < 1 or classNo > 6:
		print("invalid class number")
		sys.exit()



# it opens excel takes out the value based on classNO value from filename
def accessExcelSheet(Filename , classNo):
	workbook = load_workbook(filename = Filename)
	ws = workbook.active
	position = chr(65+classNo)+"9"
	classId = ws[position].value
	if classId is None:
		print("no class")
		sys.exit()
	return str(classId)
	



def initializeClass():
	os.system("gnome-terminal")
	sleep(2)
	pya.write("nohup zoom")
	pya.press("enter")
	sleep(5)

def navJoin():
	for i in range(10):
		pya.press("tab")


def joinClass(classId):
	pya.press("enter")
	sleep(3)
	for i in range(2):
		pya.press("tab")

	pya.write(classId)
	pya.press("enter")


def automateZoom(classId):
	initializeClass()
	navJoin()
	joinClass(classId)


filename , classNo =  inputChecker()
isClassNoValid(classNo)
classId  = accessExcelSheet(filename,classNo)
automateZoom(classId)

